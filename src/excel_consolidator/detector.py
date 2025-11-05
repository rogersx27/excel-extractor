"""Detector de estructura de archivos Excel.

Este módulo analiza archivos Excel para determinar si tienen:
- Estructura simple: Una tabla con un encabezado
- Estructura compleja: Múltiples tablas apiladas con encabezados repetidos
"""

from pathlib import Path
from typing import Dict, List, Tuple

from openpyxl import load_workbook

from logger import setup_logger

from .utils import count_non_empty_cells, is_empty_row, is_likely_header

# Logger Nivel 3 - Procesador: INFO solo archivo, detecta estructuras sin saturar consola
logger = setup_logger(
    __name__,
    level="INFO",
    console_output=False,
    file_output=True
)


class StructureType:
    """Tipos de estructura de Excel."""

    SIMPLE = "simple"
    COMPLEX = "complex"
    COMPLEX_FECHA = "complex_fecha"  # Patrón con bloques FECHA:
    UNKNOWN = "unknown"


def detect_structure(file_path: Path, sheet_name: str = None) -> Dict:
    """Detecta la estructura de un archivo Excel.

    Args:
        file_path: Ruta al archivo Excel
        sheet_name: Nombre de la hoja a analizar (None = primera hoja)

    Returns:
        Diccionario con información de la estructura:
        {
            'type': 'simple' | 'complex' | 'unknown',
            'header_rows': [list of row indices],
            'data_ranges': [(start_row, end_row), ...],
            'total_rows': int,
            'sheet_name': str
        }

    Example:
        >>> info = detect_structure(Path('datos.xlsx'))
        >>> if info['type'] == 'complex':
        ...     print(f"Encontradas {len(info['header_rows'])} tablas")
    """
    logger.info(f"Analizando estructura de: {file_path}")

    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)

        # Seleccionar hoja
        if sheet_name:
            ws = wb[sheet_name]
        else:
            ws = wb.active

        sheet_name = ws.title
        logger.debug(f"Analizando hoja: {sheet_name}")

        # Analizar filas
        header_rows = []
        data_ranges = []
        all_rows = []

        for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            all_rows.append(row)

        total_rows = len(all_rows)

        # PRIORIDAD 1: Buscar patrón FECHA: (más específico y confiable)
        fecha_rows = find_fecha_rows(all_rows)

        if fecha_rows:
            # Usar patrón FECHA como anclas
            logger.info(f"Patrón FECHA detectado: {len(fecha_rows)} bloques encontrados")
            header_rows = [fr + 1 for fr in fecha_rows]  # Header está después de FECHA
            data_ranges = determine_data_ranges_fecha(all_rows, fecha_rows)
            structure_type = StructureType.COMPLEX_FECHA
        else:
            # PRIORIDAD 2: Buscar encabezados genéricos (método anterior)
            header_rows = find_header_rows(all_rows)
            data_ranges = determine_data_ranges(all_rows, header_rows)
            structure_type = classify_structure(header_rows, data_ranges)

        wb.close()

        result = {
            "type": structure_type,
            "header_rows": header_rows,
            "data_ranges": data_ranges,
            "total_rows": total_rows,
            "sheet_name": sheet_name,
            "fecha_rows": fecha_rows if fecha_rows else [],  # Anclas FECHA
        }

        logger.info(
            f"Estructura detectada: {structure_type.upper()} - "
            f"{len(header_rows)} encabezado(s), {total_rows} filas totales"
        )

        return result

    except Exception as e:
        logger.error(f"Error detectando estructura de {file_path}: {e}")
        raise


def find_fecha_rows(rows: List[tuple]) -> List[int]:
    """Encuentra filas que contienen patrones de FECHA como ancla de bloques.

    Detecta dos patrones:
    1. "FECHA: DD/MM/YYYY ..." en columna A
    2. "DIA/ FECHA" en columna A (con fecha en columnas adyacentes)

    Args:
        rows: Lista de tuplas con valores de filas

    Returns:
        Lista de índices (1-based) de filas que contienen patrones FECHA

    Example:
        >>> rows = [(None,), ('FECHA: 09/10/2023 Lunes SWX113',), ('DIA/ FECHA', 'LUNES 22/07/2024')]
        >>> find_fecha_rows(rows)
        [2, 3]
    """
    import re

    fecha_rows = []

    for idx, row in enumerate(rows, start=1):
        if row and len(row) > 0:
            first_cell = str(row[0]).strip() if row[0] is not None else ""
            first_cell_upper = first_cell.upper()

            # PATRÓN 1: "FECHA: DD/MM/YYYY ..." en columna A
            if first_cell_upper.startswith("FECHA:"):
                fecha_rows.append(idx)
                logger.debug(f"Bloque FECHA (patrón 1) en fila {idx}: {first_cell[:50]}")
                continue

            # PATRÓN 2: "DIA/ FECHA" o "DIA / FECHA" o "DIA/FECHA" en columna A
            # Buscar si contiene ambas palabras "DIA" y "FECHA"
            if "DIA" in first_cell_upper and "FECHA" in first_cell_upper:
                # Verificar que hay una fecha en columnas adyacentes
                has_date_nearby = False
                for cell in row[1:5]:  # Revisar las siguientes 4 columnas
                    if cell:
                        cell_str = str(cell).strip()
                        # Buscar patrón de fecha DD/MM/YYYY
                        if re.search(r"\d{1,2}/\d{1,2}/\d{4}", cell_str):
                            has_date_nearby = True
                            break

                if has_date_nearby:
                    fecha_rows.append(idx)
                    logger.debug(f"Bloque FECHA (patrón 2) en fila {idx}: {first_cell}")

    return fecha_rows


def find_header_rows(rows: List[tuple]) -> List[int]:
    """Encuentra las filas que son encabezados.

    Args:
        rows: Lista de tuplas con valores de filas

    Returns:
        Lista de índices (1-based) de filas que son encabezados

    Example:
        >>> rows = [(None,), ('NOMBRE', 'DIRECCIÓN'), ('Juan', 'Calle 1')]
        >>> find_header_rows(rows)
        [2]
    """
    header_rows = []

    for idx, row in enumerate(rows, start=1):
        row_list = list(row)

        # Saltar filas vacías
        if is_empty_row(row_list):
            continue

        # Verificar si es encabezado
        if is_likely_header(row_list):
            header_rows.append(idx)
            logger.debug(f"Encabezado encontrado en fila {idx}")

    return header_rows


def determine_data_ranges_fecha(
    rows: List[tuple], fecha_rows: List[int]
) -> List[Tuple[int, int]]:
    """Determina rangos de datos basados en filas FECHA como anclas.

    Args:
        rows: Lista de tuplas con valores de filas
        fecha_rows: Índices de filas con "FECHA:"

    Returns:
        Lista de tuplas (start_row, end_row) con rangos de datos

    Example:
        >>> fecha_rows = [5, 20, 35]
        >>> determine_data_ranges_fecha(rows, fecha_rows)
        [(7, 18), (22, 33), (37, 50)]
    """
    data_ranges = []

    for i, fecha_idx in enumerate(fecha_rows):
        # Datos comienzan 2 filas después de FECHA (FECHA + 1 = header, FECHA + 2 = datos)
        start_row = fecha_idx + 2

        # Determinar fin del rango
        if i < len(fecha_rows) - 1:
            # Hay otro bloque FECHA después, buscar hasta la fila anterior
            next_fecha = fecha_rows[i + 1]
            end_row = next_fecha - 1
        else:
            # Último bloque FECHA, ir hasta el final
            end_row = len(rows)

        # Ajustar end_row para excluir filas vacías al final
        while end_row > start_row:
            if not is_empty_row(list(rows[end_row - 1])):
                break
            end_row -= 1

        if start_row <= end_row:
            data_ranges.append((start_row, end_row))
            logger.debug(
                f"Bloque FECHA {i+1}: fila FECHA={fecha_idx}, "
                f"datos={start_row}-{end_row}"
            )

    return data_ranges


def determine_data_ranges(
    rows: List[tuple], header_rows: List[int]
) -> List[Tuple[int, int]]:
    """Determina los rangos de filas que contienen datos.

    Args:
        rows: Lista de tuplas con valores de filas
        header_rows: Índices de filas de encabezados

    Returns:
        Lista de tuplas (start_row, end_row) con rangos de datos

    Example:
        >>> rows = [(...), (...), ...]
        >>> header_rows = [2, 10]
        >>> determine_data_ranges(rows, header_rows)
        [(3, 9), (11, 15)]
    """
    if not header_rows:
        # Si no hay encabezados detectados, asumir que todo es data
        # excepto las primeras filas vacías
        first_data_row = 1
        for idx, row in enumerate(rows, start=1):
            if not is_empty_row(list(row)):
                first_data_row = idx
                break

        return [(first_data_row, len(rows))]

    data_ranges = []

    for i, header_idx in enumerate(header_rows):
        start_row = header_idx + 1

        # Determinar fin del rango
        if i < len(header_rows) - 1:
            # Hay otro encabezado después
            next_header = header_rows[i + 1]
            end_row = next_header - 1
        else:
            # Último encabezado, ir hasta el final
            end_row = len(rows)

        # Ajustar end_row para excluir filas vacías al final
        while end_row > start_row:
            if not is_empty_row(list(rows[end_row - 1])):
                break
            end_row -= 1

        if start_row <= end_row:
            data_ranges.append((start_row, end_row))
            logger.debug(f"Rango de datos: filas {start_row} a {end_row}")

    return data_ranges


def classify_structure(
    header_rows: List[int], data_ranges: List[Tuple[int, int]]
) -> str:
    """Clasifica la estructura del archivo.

    Args:
        header_rows: Índices de encabezados
        data_ranges: Rangos de datos

    Returns:
        Tipo de estructura: 'simple', 'complex', o 'unknown'

    Example:
        >>> classify_structure([2], [(3, 10)])
        'simple'
        >>> classify_structure([2, 15, 28], [(3, 14), (16, 27), (29, 40)])
        'complex'
    """
    if not header_rows:
        return StructureType.UNKNOWN

    if len(header_rows) == 1:
        return StructureType.SIMPLE

    if len(header_rows) > 1:
        return StructureType.COMPLEX

    return StructureType.UNKNOWN


def get_sheet_names(file_path: Path) -> List[str]:
    """Obtiene los nombres de todas las hojas de un archivo Excel.

    Args:
        file_path: Ruta al archivo Excel

    Returns:
        Lista con nombres de hojas

    Example:
        >>> sheets = get_sheet_names(Path('datos.xlsx'))
        >>> print(sheets)
        ['Hoja1', 'Hoja2', 'Resumen']
    """
    try:
        wb = load_workbook(file_path, read_only=True)
        sheet_names = wb.sheetnames
        wb.close()
        return sheet_names
    except Exception as e:
        logger.error(f"Error obteniendo hojas de {file_path}: {e}")
        raise


def analyze_file_completely(file_path: Path) -> Dict:
    """Analiza completamente un archivo Excel (todas sus hojas).

    Args:
        file_path: Ruta al archivo Excel

    Returns:
        Diccionario con análisis de todas las hojas:
        {
            'file_path': Path,
            'total_sheets': int,
            'sheets': {
                'sheet_name': {...estructura...}
            }
        }

    Example:
        >>> analysis = analyze_file_completely(Path('datos.xlsx'))
        >>> for sheet, info in analysis['sheets'].items():
        ...     print(f"{sheet}: {info['type']}")
    """
    logger.info(f"Análisis completo de archivo: {file_path}")

    try:
        sheet_names = get_sheet_names(file_path)
        sheets_info = {}

        for sheet in sheet_names:
            sheets_info[sheet] = detect_structure(file_path, sheet)

        result = {
            "file_path": file_path,
            "total_sheets": len(sheet_names),
            "sheets": sheets_info,
        }

        logger.info(f"Análisis completado: {len(sheet_names)} hojas procesadas")

        return result

    except Exception as e:
        logger.error(f"Error en análisis completo de {file_path}: {e}")
        raise
