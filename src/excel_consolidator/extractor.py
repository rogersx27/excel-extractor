"""Extractor de datos de archivos Excel.

Este módulo extrae datos de archivos Excel según su estructura:
- Simples: Lectura directa con pandas
- Complejos: Extracción de múltiples tablas y consolidación
"""

from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook

from logger import setup_logger, setup_processor_logger

from .detector import StructureType, detect_structure
from .utils import clean_dataframe, is_empty_row

# Logger Nivel 3 - Procesador: Configuración dinámica desde variables de entorno
logger = setup_processor_logger(setup_logger, __name__)


def extract_data(
    file_path: Path,
    sheet_name: Optional[str] = None,
    structure_info: Optional[Dict] = None,
) -> pd.DataFrame:
    """Extrae datos de un archivo Excel automáticamente.

    Detecta la estructura y aplica el método de extracción apropiado.

    Args:
        file_path: Ruta al archivo Excel
        sheet_name: Nombre de la hoja (None = primera hoja)
        structure_info: Info de estructura pre-calculada (opcional)

    Returns:
        DataFrame con los datos consolidados

    Example:
        >>> df = extract_data(Path('datos.xlsx'))
        >>> print(f"Extraídas {len(df)} filas")
    """
    logger.info(f"Extrayendo datos de: {file_path}")

    # Detectar estructura si no se proporcionó
    if structure_info is None:
        structure_info = detect_structure(file_path, sheet_name)

    structure_type = structure_info["type"]

    # Usar el sheet_name de structure_info si no se proporcionó uno
    if sheet_name is None and "sheet_name" in structure_info:
        sheet_name = structure_info["sheet_name"]

    # Seleccionar método de extracción
    if structure_type == StructureType.SIMPLE:
        df = extract_simple_table(file_path, sheet_name, structure_info)
    elif structure_type == StructureType.COMPLEX_FECHA:
        df = extract_fecha_pattern_tables(file_path, sheet_name, structure_info)
    elif structure_type == StructureType.COMPLEX:
        df = extract_complex_tables(file_path, sheet_name, structure_info)
    else:
        # Estructura desconocida, intentar lectura simple
        logger.warning(
            f"Estructura desconocida en {file_path}, " f"intentando lectura simple"
        )
        df = extract_simple_table(file_path, sheet_name, structure_info)

    # Limpiar DataFrame
    df = clean_dataframe(df)

    logger.info(f"Extracción completada: {len(df)} filas, {len(df.columns)} columnas")

    return df


def extract_simple_table(
    file_path: Path,
    sheet_name: Optional[str] = None,
    structure_info: Optional[Dict] = None,
) -> pd.DataFrame:
    """Extrae una tabla simple (un solo encabezado).

    Args:
        file_path: Ruta al archivo Excel
        sheet_name: Nombre de la hoja
        structure_info: Información de estructura (opcional)

    Returns:
        DataFrame con los datos

    Example:
        >>> df = extract_simple_table(Path('clientes.xlsx'))
    """
    logger.info(f"Extrayendo tabla simple de: {file_path}")

    try:
        # Si tenemos info de estructura, usar para skiprows
        skiprows = 0
        if structure_info and structure_info["header_rows"]:
            # El encabezado está en la primera posición detectada
            header_row = structure_info["header_rows"][0]
            skiprows = header_row - 1  # pandas usa 0-indexed

        df = pd.read_excel(file_path, sheet_name=sheet_name, skiprows=skiprows)

        # Validar que obtuvimos un DataFrame (no un dict)
        if isinstance(df, dict):
            # Si es dict, tomar la primera hoja
            if len(df) > 0:
                df = list(df.values())[0]
            else:
                df = pd.DataFrame()

        logger.debug(f"Tabla simple extraída: {len(df)} filas")
        return df

    except Exception as e:
        logger.error(f"Error extrayendo tabla simple de {file_path}: {e}")
        raise


def extract_fecha_pattern_tables(
    file_path: Path,
    sheet_name: Optional[str] = None,
    structure_info: Optional[Dict] = None,
) -> pd.DataFrame:
    """Extrae múltiples bloques con patrón FECHA y agrega metadata.

    Args:
        file_path: Ruta al archivo Excel
        sheet_name: Nombre de la hoja
        structure_info: Información de estructura (requerida)

    Returns:
        DataFrame con todas las tablas consolidadas y columnas de metadata

    Example:
        >>> info = detect_structure(Path('rutas.xlsx'))
        >>> df = extract_fecha_pattern_tables(Path('rutas.xlsx'), structure_info=info)
    """
    import re

    logger.info(f"Extrayendo bloques con patrón FECHA de: {file_path}")

    if not structure_info:
        raise ValueError("structure_info es requerido para extraer patrón FECHA")

    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)

        # Seleccionar hoja
        if sheet_name:
            ws = wb[sheet_name]
        else:
            ws = wb.active

        # Leer todas las filas
        all_rows = list(ws.iter_rows(values_only=True))

        # Extraer cada bloque
        tables = []
        fecha_rows = structure_info["fecha_rows"]
        data_ranges = structure_info["data_ranges"]

        logger.info(f"Procesando {len(fecha_rows)} bloques FECHA")

        for i, (fecha_idx, (start_row, end_row)) in enumerate(
            zip(fecha_rows, data_ranges), start=1
        ):
            logger.debug(f"Bloque {i}: FECHA fila={fecha_idx}, datos={start_row}-{end_row}")

            # 1. EXTRAER METADATA de la fila FECHA
            fecha_row = all_rows[fecha_idx - 1]

            # Parsear metadata (pasar toda la fila para detectar ambos patrones)
            metadata = parse_fecha_metadata(fecha_row)

            # 2. EXTRAER ENCABEZADO (fila después de FECHA)
            header_idx = fecha_idx  # fecha_idx + 1 - 1 (convertir a 0-indexed)
            header = list(all_rows[header_idx])

            # Limpiar encabezado
            header = [
                str(col).strip() if col is not None else f"Column_{idx}"
                for idx, col in enumerate(header)
            ]

            # 3. EXTRAER DATOS
            data_rows = []
            for row_idx in range(start_row - 1, end_row):
                if row_idx < len(all_rows):
                    row = list(all_rows[row_idx])
                    if not is_empty_row(row):
                        data_rows.append(row)

            if data_rows:
                # Crear DataFrame para este bloque
                df_table = pd.DataFrame(data_rows, columns=header[: len(data_rows[0])])

                # Agregar columnas de metadata al principio
                df_table.insert(0, "fecha", metadata["fecha"])
                df_table.insert(1, "dia", metadata["dia"])
                df_table.insert(2, "vehiculo", metadata["vehiculo"])

                tables.append(df_table)
                logger.debug(
                    f"Bloque {i} extraído: {len(df_table)} filas, "
                    f"fecha={metadata['fecha']}, vehiculo={metadata['vehiculo']}"
                )

        wb.close()

        # Consolidar todas las tablas
        if not tables:
            logger.warning("No se extrajeron bloques FECHA, retornando DataFrame vacío")
            return pd.DataFrame()

        df_consolidated = pd.concat(tables, ignore_index=True, sort=False)

        logger.info(
            f"Bloques FECHA consolidados: {len(tables)} bloques → "
            f"{len(df_consolidated)} filas totales"
        )

        return df_consolidated

    except Exception as e:
        logger.error(f"Error extrayendo bloques FECHA de {file_path}: {e}")
        raise


def parse_fecha_metadata(fecha_row: tuple) -> Dict[str, str]:
    """Parsea la metadata de una fila FECHA.

    Soporta dos patrones:
    1. "FECHA: 09/10/2023 Lunes SWX113" en columna A
    2. "DIA/ FECHA" en columna A, "LUNES 22/07/2024" en columnas adyacentes

    Args:
        fecha_row: Tupla con valores de la fila FECHA

    Returns:
        Dict con 'fecha', 'dia', 'vehiculo'

    Example:
        >>> parse_fecha_metadata(("FECHA: 09/10/2023 Lunes SWX113",))
        {'fecha': '09/10/2023', 'dia': 'Lunes', 'vehiculo': 'SWX113'}
        >>> parse_fecha_metadata(("DIA/ FECHA", "LUNES 22/07/2024"))
        {'fecha': '22/07/2024', 'dia': 'LUNES', 'vehiculo': ''}
    """
    import re

    metadata = {"fecha": "", "dia": "", "vehiculo": ""}

    if not fecha_row or len(fecha_row) == 0:
        return metadata

    first_cell = str(fecha_row[0]).strip() if fecha_row[0] is not None else ""
    first_cell_upper = first_cell.upper()

    # PATRÓN 1: "FECHA: DD/MM/YYYY DayName VEHICLE" en columna A
    if first_cell_upper.startswith("FECHA:"):
        # Extraer fecha con regex: DD/MM/YYYY
        fecha_match = re.search(r"(\d{1,2}/\d{1,2}/\d{4})", first_cell)
        if fecha_match:
            metadata["fecha"] = fecha_match.group(1)

        # Extraer día y vehículo
        remaining = first_cell.upper().replace("FECHA:", "").strip()
        if metadata["fecha"]:
            remaining = remaining.replace(metadata["fecha"], "").strip()

        # Lo que queda: "- Lunes SWX113" o "Lunes SWX113"
        remaining = remaining.lstrip("-").strip()

        # Buscar códigos de vehículo conocidos
        vehicle_match = re.search(
            r"(SWX[-\s]?113|TLR[-\s]?886|SNY[-\s]?928|[A-Z]{3}[-\s]?\d{3})", remaining
        )
        if vehicle_match:
            metadata["vehiculo"] = vehicle_match.group(1).replace(" ", "")
            # El día está entre la fecha y el vehículo
            dia_text = remaining[: vehicle_match.start()].strip()
            if dia_text:
                metadata["dia"] = dia_text

    # PATRÓN 2: "DIA/ FECHA" en columna A, fecha en columnas adyacentes
    elif "DIA" in first_cell_upper and "FECHA" in first_cell_upper:
        # Buscar fecha y día en las siguientes columnas
        for cell in fecha_row[1:5]:
            if cell:
                cell_str = str(cell).strip()

                # Buscar fecha DD/MM/YYYY
                fecha_match = re.search(r"(\d{1,2}/\d{1,2}/\d{4})", cell_str)
                if fecha_match and not metadata["fecha"]:
                    metadata["fecha"] = fecha_match.group(1)

                # Extraer día (palabra antes de la fecha)
                # Ejemplo: "LUNES 22/07/2024"
                if fecha_match:
                    dia_text = cell_str[: fecha_match.start()].strip().upper()
                    if dia_text and not metadata["dia"]:
                        metadata["dia"] = dia_text

                # Buscar vehículo
                vehicle_match = re.search(
                    r"(SWX[-\s]?113|TLR[-\s]?886|SNY[-\s]?928|[A-Z]{3}[-\s]?\d{3})",
                    cell_str,
                )
                if vehicle_match and not metadata["vehiculo"]:
                    metadata["vehiculo"] = vehicle_match.group(1).replace(" ", "")

    logger.debug(f"Metadata parseada: {metadata}")
    return metadata


def extract_complex_tables(
    file_path: Path,
    sheet_name: Optional[str] = None,
    structure_info: Optional[Dict] = None,
) -> pd.DataFrame:
    """Extrae múltiples tablas apiladas y las consolida.

    Args:
        file_path: Ruta al archivo Excel
        sheet_name: Nombre de la hoja
        structure_info: Información de estructura (requerida)

    Returns:
        DataFrame con todas las tablas consolidadas

    Example:
        >>> info = detect_structure(Path('rutas.xlsx'))
        >>> df = extract_complex_tables(Path('rutas.xlsx'), structure_info=info)
    """
    logger.info(f"Extrayendo tablas complejas de: {file_path}")

    if not structure_info:
        raise ValueError("structure_info es requerido para extraer tablas complejas")

    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)

        # Seleccionar hoja
        if sheet_name:
            ws = wb[sheet_name]
        else:
            ws = wb.active

        # Leer todas las filas
        all_rows = list(ws.iter_rows(values_only=True))

        # Extraer cada tabla
        tables = []
        header_rows = structure_info["header_rows"]
        data_ranges = structure_info["data_ranges"]

        logger.info(f"Procesando {len(header_rows)} tablas")

        for i, (header_idx, (start_row, end_row)) in enumerate(
            zip(header_rows, data_ranges), start=1
        ):
            logger.debug(
                f"Tabla {i}: encabezado={header_idx}, " f"datos={start_row}-{end_row}"
            )

            # Extraer encabezado (convertir a 0-indexed)
            header = list(all_rows[header_idx - 1])

            # Limpiar encabezado (eliminar None y vacíos)
            header = [
                str(col).strip() if col is not None else f"Column_{idx}"
                for idx, col in enumerate(header)
            ]

            # Extraer datos
            data_rows = []
            for row_idx in range(start_row - 1, end_row):
                if row_idx < len(all_rows):
                    row = list(all_rows[row_idx])
                    if not is_empty_row(row):
                        data_rows.append(row)

            if data_rows:
                # Crear DataFrame para esta tabla
                df_table = pd.DataFrame(data_rows, columns=header[: len(data_rows[0])])
                tables.append(df_table)
                logger.debug(f"Tabla {i} extraída: {len(df_table)} filas")

        wb.close()

        # Consolidar todas las tablas
        if not tables:
            logger.warning("No se extrajeron tablas, retornando DataFrame vacío")
            return pd.DataFrame()

        # Intentar concatenar con align de columnas
        df_consolidated = pd.concat(tables, ignore_index=True, sort=False)

        logger.info(
            f"Tablas consolidadas: {len(tables)} tablas → "
            f"{len(df_consolidated)} filas totales"
        )

        return df_consolidated

    except Exception as e:
        logger.error(f"Error extrayendo tablas complejas de {file_path}: {e}")
        raise


def extract_range(
    file_path: Path,
    sheet_name: Optional[str],
    start_row: int,
    end_row: int,
    start_col: int = 1,
    end_col: Optional[int] = None,
) -> pd.DataFrame:
    """Extrae un rango específico de celdas como DataFrame.

    Args:
        file_path: Ruta al archivo Excel
        sheet_name: Nombre de la hoja
        start_row: Fila inicial (1-indexed)
        end_row: Fila final (1-indexed)
        start_col: Columna inicial (1-indexed, default=1)
        end_col: Columna final (1-indexed, None=todas)

    Returns:
        DataFrame con los datos del rango

    Example:
        >>> df = extract_range(Path('datos.xlsx'), 'Hoja1', 5, 20)
    """
    logger.info(
        f"Extrayendo rango de {file_path}: "
        f"filas {start_row}-{end_row}, cols {start_col}-{end_col or 'all'}"
    )

    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)

        if sheet_name:
            ws = wb[sheet_name]
        else:
            ws = wb.active

        # Leer rango
        rows_data = []
        for row in ws.iter_rows(
            min_row=start_row,
            max_row=end_row,
            min_col=start_col,
            max_col=end_col,
            values_only=True,
        ):
            rows_data.append(list(row))

        wb.close()

        # Crear DataFrame (primera fila como encabezado)
        if len(rows_data) > 1:
            df = pd.DataFrame(rows_data[1:], columns=rows_data[0])
        else:
            df = pd.DataFrame(rows_data)

        logger.info(f"Rango extraído: {len(df)} filas")
        return df

    except Exception as e:
        logger.error(f"Error extrayendo rango de {file_path}: {e}")
        raise


def extract_all_sheets(file_path: Path) -> Dict[str, pd.DataFrame]:
    """Extrae datos de todas las hojas de un archivo Excel.

    Args:
        file_path: Ruta al archivo Excel

    Returns:
        Diccionario {nombre_hoja: DataFrame}

    Example:
        >>> sheets_data = extract_all_sheets(Path('workbook.xlsx'))
        >>> for sheet, df in sheets_data.items():
        ...     print(f"{sheet}: {len(df)} filas")
    """
    logger.info(f"Extrayendo todas las hojas de: {file_path}")

    try:
        # Obtener nombres de hojas
        wb = load_workbook(file_path, read_only=True)
        sheet_names = wb.sheetnames
        wb.close()

        # Extraer cada hoja
        sheets_data = {}
        for sheet in sheet_names:
            logger.debug(f"Procesando hoja: {sheet}")
            df = extract_data(file_path, sheet_name=sheet)
            sheets_data[sheet] = df

        logger.info(f"Extraídas {len(sheets_data)} hojas")
        return sheets_data

    except Exception as e:
        logger.error(f"Error extrayendo hojas de {file_path}: {e}")
        raise


def preview_data(
    file_path: Path, sheet_name: Optional[str] = None, n_rows: int = 10
) -> pd.DataFrame:
    """Obtiene una vista previa de los datos (primeras n filas).

    Args:
        file_path: Ruta al archivo Excel
        sheet_name: Nombre de la hoja
        n_rows: Número de filas a mostrar (default=10)

    Returns:
        DataFrame con las primeras n filas

    Example:
        >>> preview = preview_data(Path('datos.xlsx'), n_rows=5)
        >>> print(preview)
    """
    logger.info(f"Generando preview de: {file_path}")

    try:
        df = extract_data(file_path, sheet_name)
        preview_df = df.head(n_rows)

        logger.info(f"Preview generado: {len(preview_df)} filas de {len(df)} totales")

        return preview_df

    except Exception as e:
        logger.error(f"Error generando preview de {file_path}: {e}")
        raise
