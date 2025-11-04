"""Manejador avanzado de Excel para manipulación multi-hoja.

Este módulo proporciona la clase ExcelHandler para operaciones complejas
que requieren mantener archivos abiertos y gestionar múltiples hojas.

Características:
- Context manager para gestión automática de recursos
- Lazy loading del workbook
- Manipulación de hojas (agregar, eliminar, renombrar)
- Extracción de rangos específicos
- Formato básico con openpyxl

Use ExcelHandler para:
- Manipular múltiples hojas
- Modificar archivos existentes
- Operaciones complejas que requieren varias acciones

Use QuickExcel (quick.py) para:
- Lectura/escritura simple de una línea
- Scripts rápidos sin gestión de estado
"""
from pathlib import Path
from typing import Any, List, Optional, Union

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from logger import setup_logger
from .exceptions import (
    FileOperationError,
    InvalidFileFormatError,
    SheetAlreadyExistsError,
    SheetNotFoundError,
)

logger = setup_logger(__name__)


class ExcelHandler:
    """Manejador avanzado para operaciones complejas con archivos Excel.

    Esta clase mantiene un archivo Excel abierto para realizar múltiples
    operaciones. Implementa context manager para garantizar el cierre
    automático del archivo.

    Example:
        >>> # Uso con context manager (recomendado)
        >>> with ExcelHandler('reporte.xlsx') as excel:
        ...     excel.add_sheet('Ventas', df_ventas)
        ...     excel.add_sheet('Resumen', df_resumen)
        ...     excel.rename_sheet('Sheet1', 'Principal')
        ...     excel.save()
        ... # Auto-close garantizado
        >>>
        >>> # Uso sin context manager (manual)
        >>> handler = ExcelHandler('datos.xlsx')
        >>> handler.add_sheet('Nueva', df)
        >>> handler.save()
        >>> handler.close()  # ¡No olvidar!
    """

    # Extensiones soportadas
    VALID_EXTENSIONS = {'.xlsx', '.xlsm', '.xltx', '.xltm'}

    def __init__(self, file_path: Union[str, Path], validate: bool = True):
        """Inicializa el manejador de Excel.

        Args:
            file_path: Ruta al archivo Excel (existente o nuevo)
            validate: Si True, valida extensión del archivo (default: True)

        Raises:
            InvalidFileFormatError: Si la extensión no es válida
        """
        self.file_path = Path(file_path)
        self._workbook: Optional[Workbook] = None

        if validate:
            self._validate_file()

    def _validate_file(self):
        """Valida que la extensión del archivo sea correcta.

        Raises:
            InvalidFileFormatError: Si la extensión no es válida
        """
        if self.file_path.suffix.lower() not in ExcelHandler.VALID_EXTENSIONS:
            raise InvalidFileFormatError(
                f"Extensión no soportada: {self.file_path.suffix}. "
                f"Extensiones válidas: {', '.join(ExcelHandler.VALID_EXTENSIONS)}"
            )

    @property
    def wb(self) -> Workbook:
        """Acceso lazy al workbook (carga automáticamente si es None).

        Returns:
            Workbook de openpyxl

        Raises:
            FileOperationError: Si hay error al cargar el archivo

        Example:
            >>> handler = ExcelHandler('datos.xlsx')
            >>> # El workbook se carga automáticamente en el primer acceso
            >>> print(handler.wb.sheetnames)
        """
        if self._workbook is None:
            self._load_workbook()
        return self._workbook

    def _load_workbook(self):
        """Carga el workbook desde el archivo.

        Si el archivo no existe, crea un nuevo workbook.

        Raises:
            FileOperationError: Si hay error al cargar el archivo
        """
        try:
            if self.file_path.exists():
                logger.info(f"Cargando Excel existente: {self.file_path}")
                self._workbook = load_workbook(self.file_path)
            else:
                logger.info(f"Creando nuevo Excel: {self.file_path}")
                self._workbook = Workbook()
        except Exception as e:
            logger.error(f"Error cargando workbook {self.file_path}: {e}")
            raise FileOperationError(
                f"No se pudo cargar el archivo {self.file_path}: {e}"
            ) from e

    # ==================== CONTEXT MANAGER ====================

    def __enter__(self):
        """Permite usar with statement.

        Returns:
            Self para encadenamiento

        Example:
            >>> with ExcelHandler('datos.xlsx') as excel:
            ...     excel.add_sheet('Nueva')
        """
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        """Cierra automáticamente el workbook al salir del contexto.

        Args:
            exc_type: Tipo de excepción (si hubo)
            exc_val: Valor de la excepción
            exc_tb: Traceback de la excepción

        Returns:
            False para propagar excepciones
        """
        self.close()
        return False

    # ==================== INFORMACIÓN ====================

    @property
    def sheet_names(self) -> List[str]:
        """Obtiene los nombres de todas las hojas del archivo.

        Returns:
            Lista con los nombres de las hojas

        Example:
            >>> handler = ExcelHandler('datos.xlsx')
            >>> print(handler.sheet_names)
            ['Ventas', 'Clientes', 'Productos']
        """
        return self.wb.sheetnames

    def sheet_exists(self, sheet_name: str) -> bool:
        """Verifica si una hoja existe en el workbook.

        Args:
            sheet_name: Nombre de la hoja a verificar

        Returns:
            True si la hoja existe, False en caso contrario

        Example:
            >>> handler = ExcelHandler('datos.xlsx')
            >>> if handler.sheet_exists('Ventas'):
            ...     print('La hoja existe')
        """
        return sheet_name in self.wb.sheetnames

    # ==================== MANIPULACIÓN DE HOJAS ====================

    def add_sheet(
        self,
        sheet_name: str,
        data: Optional[pd.DataFrame] = None,
        index: int = None,
        overwrite: bool = False
    ):
        """Añade una nueva hoja al Excel.

        Args:
            sheet_name: Nombre de la nueva hoja
            data: DataFrame opcional con datos para la hoja
            index: Índice donde insertar la hoja (None = al final)
            overwrite: Si True, sobrescribe hoja existente (default: False)

        Raises:
            SheetAlreadyExistsError: Si la hoja ya existe y overwrite=False

        Example:
            >>> with ExcelHandler('reporte.xlsx') as excel:
            ...     df = pd.DataFrame({'A': [1, 2, 3]})
            ...     excel.add_sheet('NuevaHoja', df)
            ...     excel.save()
        """
        # Verificar si la hoja ya existe
        if self.sheet_exists(sheet_name):
            if not overwrite:
                raise SheetAlreadyExistsError(
                    f"La hoja '{sheet_name}' ya existe. "
                    f"Use overwrite=True para sobrescribirla."
                )
            else:
                logger.info(f"Sobrescribiendo hoja existente: '{sheet_name}'")
                self.delete_sheet(sheet_name)

        # Crear nueva hoja
        ws = self.wb.create_sheet(sheet_name, index=index)

        # Agregar datos si se proporcionan
        if data is not None and not data.empty:
            for r_idx, row in enumerate(dataframe_to_rows(data, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

        logger.info(f"Hoja '{sheet_name}' añadida exitosamente")

    def delete_sheet(self, sheet_name: str):
        """Elimina una hoja del Excel.

        Args:
            sheet_name: Nombre de la hoja a eliminar

        Raises:
            SheetNotFoundError: Si la hoja no existe

        Example:
            >>> with ExcelHandler('datos.xlsx') as excel:
            ...     excel.delete_sheet('HojaVieja')
            ...     excel.save()
        """
        if not self.sheet_exists(sheet_name):
            raise SheetNotFoundError(
                f"Hoja '{sheet_name}' no encontrada. "
                f"Hojas disponibles: {', '.join(self.sheet_names)}"
            )

        del self.wb[sheet_name]
        logger.info(f"Hoja '{sheet_name}' eliminada")

    def rename_sheet(self, old_name: str, new_name: str):
        """Renombra una hoja del Excel.

        Args:
            old_name: Nombre actual de la hoja
            new_name: Nuevo nombre para la hoja

        Raises:
            SheetNotFoundError: Si la hoja no existe
            SheetAlreadyExistsError: Si el nuevo nombre ya existe

        Example:
            >>> with ExcelHandler('datos.xlsx') as excel:
            ...     excel.rename_sheet('Sheet1', 'Ventas')
            ...     excel.save()
        """
        if not self.sheet_exists(old_name):
            raise SheetNotFoundError(
                f"Hoja '{old_name}' no encontrada. "
                f"Hojas disponibles: {', '.join(self.sheet_names)}"
            )

        if self.sheet_exists(new_name):
            raise SheetAlreadyExistsError(
                f"Ya existe una hoja con el nombre '{new_name}'"
            )

        self.wb[old_name].title = new_name
        logger.info(f"Hoja renombrada: '{old_name}' -> '{new_name}'")

    # ==================== EXTRACCIÓN DE DATOS ====================

    def extract_sheet_to_dataframe(
        self,
        sheet_name: Union[str, int],
        **kwargs
    ) -> pd.DataFrame:
        """Extrae una hoja específica como DataFrame.

        Args:
            sheet_name: Nombre o índice de la hoja
            **kwargs: Argumentos adicionales para pd.read_excel

        Returns:
            DataFrame con los datos de la hoja

        Raises:
            SheetNotFoundError: Si la hoja no existe (cuando es string)

        Example:
            >>> handler = ExcelHandler('datos.xlsx')
            >>> df = handler.extract_sheet_to_dataframe('Ventas')
            >>> print(df.head())
        """
        # Validar que la hoja existe si es string
        if isinstance(sheet_name, str) and not self.sheet_exists(sheet_name):
            raise SheetNotFoundError(
                f"Hoja '{sheet_name}' no encontrada. "
                f"Hojas disponibles: {', '.join(self.sheet_names)}"
            )

        logger.info(f"Extrayendo hoja '{sheet_name}' como DataFrame")
        return pd.read_excel(self.file_path, sheet_name=sheet_name, **kwargs)

    def extract_range(
        self,
        sheet_name: str,
        start_row: int,
        start_col: int,
        end_row: int,
        end_col: int
    ) -> List[List[Any]]:
        """Extrae un rango específico de celdas.

        Args:
            sheet_name: Nombre de la hoja
            start_row: Fila inicial (1-indexed)
            start_col: Columna inicial (1-indexed)
            end_row: Fila final (1-indexed)
            end_col: Columna final (1-indexed)

        Returns:
            Lista de listas con los valores del rango

        Raises:
            SheetNotFoundError: Si la hoja no existe

        Example:
            >>> handler = ExcelHandler('datos.xlsx')
            >>> # Extraer rango A1:C5
            >>> data = handler.extract_range('Hoja1', 1, 1, 5, 3)
            >>> print(data)
        """
        if not self.sheet_exists(sheet_name):
            raise SheetNotFoundError(
                f"Hoja '{sheet_name}' no encontrada. "
                f"Hojas disponibles: {', '.join(self.sheet_names)}"
            )

        ws = self.wb[sheet_name]
        data = []

        for row in ws.iter_rows(
            min_row=start_row,
            max_row=end_row,
            min_col=start_col,
            max_col=end_col,
            values_only=True
        ):
            data.append(list(row))

        logger.info(
            f"Rango extraído de '{sheet_name}': "
            f"{len(data)} filas × {len(data[0]) if data else 0} columnas"
        )
        return data

    # ==================== GUARDAR Y CERRAR ====================

    def save(self, file_path: Optional[Union[str, Path]] = None):
        """Guarda los cambios en el archivo Excel.

        Args:
            file_path: Ruta opcional para guardar (si es diferente a la original)

        Raises:
            FileOperationError: Si hay error al guardar

        Example:
            >>> with ExcelHandler('datos.xlsx') as excel:
            ...     excel.add_sheet('NuevaHoja')
            ...     excel.save()
            ...     # O guardar como otro archivo
            ...     excel.save('copia_datos.xlsx')
        """
        if self._workbook is None:
            logger.warning("No hay workbook cargado para guardar")
            return

        save_path = Path(file_path) if file_path else self.file_path

        # Crear directorio si no existe
        save_path.parent.mkdir(parents=True, exist_ok=True)

        try:
            self.wb.save(save_path)
            logger.info(f"Archivo guardado: {save_path}")
        except PermissionError as e:
            logger.error(f"Error de permisos al guardar {save_path}: {e}")
            raise FileOperationError(
                f"No se puede guardar el archivo {save_path}. "
                f"Puede estar abierto en Excel u otra aplicación."
            ) from e
        except Exception as e:
            logger.error(f"Error guardando {save_path}: {e}")
            raise FileOperationError(
                f"Error al guardar el archivo {save_path}: {e}"
            ) from e

    def close(self):
        """Cierra el workbook y libera recursos.

        Es importante llamar a este método cuando se termina de trabajar
        con el archivo, o usar el context manager (with statement) para
        cierre automático.

        Example:
            >>> handler = ExcelHandler('datos.xlsx')
            >>> handler.add_sheet('Nueva')
            >>> handler.save()
            >>> handler.close()  # Libera recursos
        """
        if self._workbook:
            self._workbook.close()
            self._workbook = None
            logger.info(f"Workbook cerrado: {self.file_path}")

    # ==================== MÉTODOS AUXILIARES ====================

    def __repr__(self) -> str:
        """Representación string del objeto."""
        status = "cargado" if self._workbook else "no cargado"
        return f"ExcelHandler(file='{self.file_path}', status='{status}')"

    def __str__(self) -> str:
        """String informativo del objeto."""
        if self._workbook:
            return (
                f"ExcelHandler: {self.file_path}\n"
                f"Hojas: {', '.join(self.sheet_names)}"
            )
        return f"ExcelHandler: {self.file_path} (no cargado)"
