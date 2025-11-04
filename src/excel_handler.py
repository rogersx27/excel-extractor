"""Módulo para manipulación avanzada de archivos Excel.

Este módulo proporciona funciones para leer, escribir y manipular
archivos Excel de forma robusta usando las mejores prácticas.
"""
import logging
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import xlsxwriter

logger = logging.getLogger(__name__)


class ExcelHandler:
    """Manejador principal para operaciones con archivos Excel."""

    def __init__(self, file_path: Union[str, Path]):
        """
        Inicializa el manejador de Excel.

        Args:
            file_path: Ruta al archivo Excel
        """
        self.file_path = Path(file_path)
        self.workbook: Optional[Workbook] = None

    # ==================== LECTURA ====================

    @staticmethod
    def read_excel_pandas(
        file_path: Union[str, Path],
        sheet_name: Optional[Union[str, int]] = 0,
        **kwargs
    ) -> pd.DataFrame:
        """
        Lee un archivo Excel usando pandas (rápido para análisis de datos).

        Args:
            file_path: Ruta al archivo Excel
            sheet_name: Nombre o índice de la hoja (default: primera hoja)
            **kwargs: Argumentos adicionales para pd.read_excel

        Returns:
            DataFrame con los datos del Excel

        Example:
            >>> df = ExcelHandler.read_excel_pandas('datos.xlsx', sheet_name='Ventas')
            >>> print(df.head())
        """
        logger.info(f"Leyendo Excel con pandas: {file_path}")
        return pd.read_excel(file_path, sheet_name=sheet_name, **kwargs)

    @staticmethod
    def read_all_sheets_pandas(file_path: Union[str, Path]) -> Dict[str, pd.DataFrame]:
        """
        Lee todas las hojas de un Excel en un diccionario de DataFrames.

        Args:
            file_path: Ruta al archivo Excel

        Returns:
            Diccionario con {nombre_hoja: DataFrame}

        Example:
            >>> sheets = ExcelHandler.read_all_sheets_pandas('datos.xlsx')
            >>> for name, df in sheets.items():
            ...     print(f"Hoja: {name}, Filas: {len(df)}")
        """
        logger.info(f"Leyendo todas las hojas: {file_path}")
        return pd.read_excel(file_path, sheet_name=None)

    def read_excel_openpyxl(self) -> Workbook:
        """
        Lee el archivo Excel usando openpyxl (mejor para formato y estilos).

        Returns:
            Objeto Workbook de openpyxl

        Example:
            >>> handler = ExcelHandler('datos.xlsx')
            >>> wb = handler.read_excel_openpyxl()
            >>> ws = wb['Hoja1']
            >>> print(ws['A1'].value)
        """
        logger.info(f"Leyendo Excel con openpyxl: {self.file_path}")
        self.workbook = load_workbook(self.file_path)
        return self.workbook

    def get_sheet_names(self) -> List[str]:
        """
        Obtiene los nombres de todas las hojas del archivo.

        Returns:
            Lista con los nombres de las hojas

        Example:
            >>> handler = ExcelHandler('datos.xlsx')
            >>> handler.read_excel_openpyxl()
            >>> print(handler.get_sheet_names())
        """
        if self.workbook is None:
            self.read_excel_openpyxl()
        return self.workbook.sheetnames

    # ==================== ESCRITURA ====================

    @staticmethod
    def write_dataframe_to_excel(
        df: pd.DataFrame,
        file_path: Union[str, Path],
        sheet_name: str = 'Sheet1',
        index: bool = False,
        **kwargs
    ):
        """
        Escribe un DataFrame a un archivo Excel usando pandas.

        Args:
            df: DataFrame a escribir
            file_path: Ruta del archivo de salida
            sheet_name: Nombre de la hoja
            index: Incluir el índice del DataFrame
            **kwargs: Argumentos adicionales para df.to_excel

        Example:
            >>> df = pd.DataFrame({'A': [1, 2, 3], 'B': [4, 5, 6]})
            >>> ExcelHandler.write_dataframe_to_excel(df, 'salida.xlsx', sheet_name='Datos')
        """
        logger.info(f"Escribiendo DataFrame a {file_path}")
        df.to_excel(file_path, sheet_name=sheet_name, index=index, **kwargs)

    @staticmethod
    def write_multiple_sheets(
        data: Dict[str, pd.DataFrame],
        file_path: Union[str, Path],
        index: bool = False
    ):
        """
        Escribe múltiples DataFrames en diferentes hojas de un mismo archivo.

        Args:
            data: Diccionario {nombre_hoja: DataFrame}
            file_path: Ruta del archivo de salida
            index: Incluir el índice del DataFrame

        Example:
            >>> data = {
            ...     'Ventas': df_ventas,
            ...     'Clientes': df_clientes,
            ...     'Productos': df_productos
            ... }
            >>> ExcelHandler.write_multiple_sheets(data, 'reporte.xlsx')
        """
        logger.info(f"Escribiendo {len(data)} hojas a {file_path}")
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            for sheet_name, df in data.items():
                df.to_excel(writer, sheet_name=sheet_name, index=index)

    # ==================== CREACIÓN CON FORMATO ====================

    @staticmethod
    def create_formatted_excel(
        df: pd.DataFrame,
        file_path: Union[str, Path],
        sheet_name: str = 'Sheet1',
        auto_filter: bool = True,
        freeze_panes: bool = True
    ):
        """
        Crea un Excel con formato profesional usando xlsxwriter.

        Args:
            df: DataFrame a escribir
            file_path: Ruta del archivo de salida
            sheet_name: Nombre de la hoja
            auto_filter: Activar autofiltro en encabezados
            freeze_panes: Congelar primera fila

        Example:
            >>> df = pd.DataFrame({'Nombre': ['Juan', 'María'], 'Edad': [30, 25]})
            >>> ExcelHandler.create_formatted_excel(df, 'reporte_bonito.xlsx')
        """
        logger.info(f"Creando Excel formateado: {file_path}")

        # Crear archivo con xlsxwriter
        writer = pd.ExcelWriter(file_path, engine='xlsxwriter')
        df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Obtener objetos del workbook
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]

        # Formato para encabezados
        header_format = workbook.add_format({
            'bold': True,
            'text_wrap': True,
            'valign': 'top',
            'fg_color': '#4472C4',
            'font_color': 'white',
            'border': 1
        })

        # Aplicar formato a encabezados
        for col_num, value in enumerate(df.columns.values):
            worksheet.write(0, col_num, value, header_format)
            # Ajustar ancho de columna
            column_len = max(df[value].astype(str).map(len).max(), len(value)) + 2
            worksheet.set_column(col_num, col_num, column_len)

        # Aplicar autofiltro
        if auto_filter:
            worksheet.autofilter(0, 0, len(df), len(df.columns) - 1)

        # Congelar paneles
        if freeze_panes:
            worksheet.freeze_panes(1, 0)

        writer.close()
        logger.info(f"Excel formateado creado exitosamente: {file_path}")

    # ==================== MANIPULACIÓN DE HOJAS ====================

    def add_sheet(self, sheet_name: str, data: Optional[pd.DataFrame] = None):
        """
        Añade una nueva hoja al Excel.

        Args:
            sheet_name: Nombre de la nueva hoja
            data: DataFrame opcional con datos para la hoja

        Example:
            >>> handler = ExcelHandler('datos.xlsx')
            >>> handler.read_excel_openpyxl()
            >>> df = pd.DataFrame({'A': [1, 2, 3]})
            >>> handler.add_sheet('NuevaHoja', df)
            >>> handler.save()
        """
        if self.workbook is None:
            self.read_excel_openpyxl()

        ws = self.workbook.create_sheet(sheet_name)

        if data is not None:
            for r_idx, row in enumerate(dataframe_to_rows(data, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

        logger.info(f"Hoja '{sheet_name}' añadida")

    def delete_sheet(self, sheet_name: str):
        """
        Elimina una hoja del Excel.

        Args:
            sheet_name: Nombre de la hoja a eliminar

        Example:
            >>> handler = ExcelHandler('datos.xlsx')
            >>> handler.read_excel_openpyxl()
            >>> handler.delete_sheet('HojaVieja')
            >>> handler.save()
        """
        if self.workbook is None:
            self.read_excel_openpyxl()

        if sheet_name in self.workbook.sheetnames:
            del self.workbook[sheet_name]
            logger.info(f"Hoja '{sheet_name}' eliminada")
        else:
            logger.warning(f"Hoja '{sheet_name}' no encontrada")

    def rename_sheet(self, old_name: str, new_name: str):
        """
        Renombra una hoja del Excel.

        Args:
            old_name: Nombre actual de la hoja
            new_name: Nuevo nombre para la hoja

        Example:
            >>> handler = ExcelHandler('datos.xlsx')
            >>> handler.read_excel_openpyxl()
            >>> handler.rename_sheet('Sheet1', 'Ventas')
            >>> handler.save()
        """
        if self.workbook is None:
            self.read_excel_openpyxl()

        if old_name in self.workbook.sheetnames:
            self.workbook[old_name].title = new_name
            logger.info(f"Hoja renombrada: '{old_name}' -> '{new_name}'")
        else:
            logger.warning(f"Hoja '{old_name}' no encontrada")

    # ==================== EXTRACCIÓN DE DATOS ====================

    def extract_sheet_to_dataframe(self, sheet_name: Union[str, int]) -> pd.DataFrame:
        """
        Extrae una hoja específica como DataFrame.

        Args:
            sheet_name: Nombre o índice de la hoja

        Returns:
            DataFrame con los datos de la hoja

        Example:
            >>> handler = ExcelHandler('datos.xlsx')
            >>> df = handler.extract_sheet_to_dataframe('Ventas')
            >>> print(df.head())
        """
        logger.info(f"Extrayendo hoja '{sheet_name}' como DataFrame")
        return pd.read_excel(self.file_path, sheet_name=sheet_name)

    def extract_range(
        self,
        sheet_name: str,
        start_row: int,
        start_col: int,
        end_row: int,
        end_col: int
    ) -> List[List[Any]]:
        """
        Extrae un rango específico de celdas.

        Args:
            sheet_name: Nombre de la hoja
            start_row: Fila inicial (1-indexed)
            start_col: Columna inicial (1-indexed)
            end_row: Fila final
            end_col: Columna final

        Returns:
            Lista de listas con los valores del rango

        Example:
            >>> handler = ExcelHandler('datos.xlsx')
            >>> handler.read_excel_openpyxl()
            >>> data = handler.extract_range('Hoja1', 1, 1, 5, 3)
            >>> print(data)
        """
        if self.workbook is None:
            self.read_excel_openpyxl()

        ws = self.workbook[sheet_name]
        data = []

        for row in ws.iter_rows(
            min_row=start_row,
            max_row=end_row,
            min_col=start_col,
            max_col=end_col,
            values_only=True
        ):
            data.append(list(row))

        logger.info(f"Rango extraído: {len(data)} filas")
        return data

    # ==================== GUARDAR ====================

    def save(self, file_path: Optional[Union[str, Path]] = None):
        """
        Guarda los cambios en el archivo Excel.

        Args:
            file_path: Ruta opcional para guardar (si es diferente a la original)

        Example:
            >>> handler = ExcelHandler('datos.xlsx')
            >>> handler.read_excel_openpyxl()
            >>> handler.add_sheet('NuevaHoja')
            >>> handler.save()
        """
        if self.workbook is None:
            logger.warning("No hay workbook cargado para guardar")
            return

        save_path = Path(file_path) if file_path else self.file_path
        self.workbook.save(save_path)
        logger.info(f"Archivo guardado: {save_path}")

    def close(self):
        """Cierra el workbook."""
        if self.workbook:
            self.workbook.close()
            self.workbook = None
            logger.info("Workbook cerrado")


# ==================== FUNCIONES DE UTILIDAD ====================

def merge_excel_files(
    file_paths: List[Union[str, Path]],
    output_path: Union[str, Path],
    sheet_name: Optional[str] = None
) -> pd.DataFrame:
    """
    Combina múltiples archivos Excel en uno solo.

    Args:
        file_paths: Lista de rutas de archivos a combinar
        output_path: Ruta del archivo de salida
        sheet_name: Nombre de la hoja a leer (si es None, lee la primera)

    Returns:
        DataFrame combinado

    Example:
        >>> files = ['ventas_enero.xlsx', 'ventas_febrero.xlsx', 'ventas_marzo.xlsx']
        >>> df_combined = merge_excel_files(files, 'ventas_trimestre.xlsx')
    """
    logger.info(f"Combinando {len(file_paths)} archivos Excel")

    dataframes = []
    for file_path in file_paths:
        df = pd.read_excel(file_path, sheet_name=sheet_name or 0)
        dataframes.append(df)

    combined_df = pd.concat(dataframes, ignore_index=True)
    combined_df.to_excel(output_path, index=False)

    logger.info(f"Archivos combinados guardados en: {output_path}")
    return combined_df


def split_excel_by_column(
    file_path: Union[str, Path],
    column_name: str,
    output_dir: Union[str, Path],
    sheet_name: Optional[str] = None
):
    """
    Divide un Excel en múltiples archivos basándose en los valores de una columna.

    Args:
        file_path: Ruta del archivo Excel
        column_name: Nombre de la columna para dividir
        output_dir: Directorio donde guardar los archivos
        sheet_name: Nombre de la hoja a procesar

    Example:
        >>> split_excel_by_column('ventas.xlsx', 'Región', 'salida/')
        # Crea: salida/Norte.xlsx, salida/Sur.xlsx, etc.
    """
    logger.info(f"Dividiendo Excel por columna: {column_name}")

    df = pd.read_excel(file_path, sheet_name=sheet_name or 0)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    for value in df[column_name].unique():
        filtered_df = df[df[column_name] == value]
        output_file = output_dir / f"{value}.xlsx"
        filtered_df.to_excel(output_file, index=False)
        logger.info(f"Archivo creado: {output_file} ({len(filtered_df)} filas)")
